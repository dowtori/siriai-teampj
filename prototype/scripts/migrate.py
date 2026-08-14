# -*- coding: utf-8 -*-
"""
운영 구글시트 → 정규화 데이터셋

입력 : MCP 로 내려받은 시트 텍스트 2종 (마스터시트 / 캠페인 진행시트)
출력 : prototype/data/db.json  — 향후 PostgreSQL 스키마와 같은 모양

브랜드 / 캠페인 / 캠페인참여 3계층으로 나누고, 진행시트 블록을 캠페인에 연결한다.
연결 성공률은 그대로 리포트한다 (동일인·동일캠페인 판별이 이 프로젝트 최대 리스크이므로
숨기지 않고 드러낸다).
"""
import sys, io, json, re, hashlib, pathlib, datetime
from collections import defaultdict

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

TOOL = pathlib.Path(
    r"C:\Users\WD\.claude\projects\C--Users-WD-Desktop-siriai----siriai-teampj"
    r"\121b005d-05dc-4fa5-bf66-11af67188334\tool-results"
)
PARTS  = TOOL / "mcp-caa4e976-a14d-4e0f-8b3f-2459cab8950d-read_file_content-1786585266695.txt"
# 마스터시트는 xlsx 원본을 쓴다. 텍스트 추출은 탭당 73행에서 잘려 138건 중 68건만 들어왔다.
MASTER_XLSX = pathlib.Path(__file__).resolve().parent / "master.xlsx"
# 진행시트를 되짚어 만든 협업 인플루언서 명단 (계정 · 참여 캠페인 수 · 캠페인 목록)
CREATORS_CSV = pathlib.Path(__file__).resolve().parent / "creators.csv"
# 실데이터는 db.local.json 에만 쓴다. 배포에 올라가는 db.json 은
# scripts/anonymize.py 가 가명 처리해서 만든다.
OUT    = pathlib.Path(__file__).resolve().parents[1] / "data" / "db.local.json"


# ─────────────────────────── 공통 ───────────────────────────
def load(p):
    return json.load(open(p, encoding="utf-8"))["fileContent"].split("\n")


def cells(ln):
    s = ln.strip()
    if s.startswith("|"): s = s[1:]
    if s.endswith("|"):   s = s[:-1]
    # 시트 텍스트는 마크다운 이스케이프가 섞여 있다
    return [c.strip().replace("\\_", "_").replace("\\*", "*")
             .replace("\\[", "[").replace("\\]", "]").replace("\\.", ".")
             .replace("\\|", "|").replace("\\#", "#").replace("\\-", "-")
            for c in s.split("|")]


def money(v):
    """'₩1,900,000' → 1900000. 시트에는 '936-' 같은 깨진 값도 섞여 있어 방어한다."""
    if not v: return None
    s = str(v).strip()
    neg = s.startswith("-") or s.startswith("(")
    d = re.sub(r"\D", "", s)
    if not d: return None
    return -int(d) if neg else int(d)


def pct(v):
    if not v: return None
    m = re.search(r"(-?\d+(?:\.\d+)?)\s*%", str(v))
    return round(float(m.group(1)) / 100, 4) if m else None


def num(v):
    """'86,750' → 86750"""
    if v is None or v == "": return None
    s = str(v).strip()
    neg = s.startswith("-")
    d = re.sub(r"\D", "", s.split(".")[0])
    if not d: return None
    return -int(d) if neg else int(d)


def dec(v):
    if v is None or v == "": return None
    m = re.search(r"-?\d+(?:\.\d+)?", str(v).replace(",", ""))
    return float(m.group(0)) if m else None


def kdate(v):
    """'26년 2월 8일' → '2026-02-08'"""
    if not v: return None
    m = re.match(r"(\d{2})년\s*(\d{1,2})월\s*(\d{1,2})일", v.strip())
    if m:
        y, mo, d = m.groups()
        return f"20{y}-{int(mo):02d}-{int(d):02d}"
    return None


def kmonth(v):
    if not v: return None
    m = re.match(r"(\d{2})년\s*(\d{1,2})월", v.strip())
    return f"20{m.group(1)}-{int(m.group(2)):02d}" if m else None


def sid(*parts):
    return hashlib.sha1("|".join(str(p) for p in parts).encode()).hexdigest()[:10]


def norm_url(u):
    """계정 주소 형식 통일 — 동일인 판별의 유일한 기준"""
    if not u: return None
    u = u.strip().split("?")[0].rstrip("/")
    u = re.sub(r"^https?://", "", u, flags=re.I)
    u = re.sub(r"^www\.", "", u, flags=re.I)
    return u.lower() or None


NOISE = {":-:", "", "▼", "상태", "캠페인명", "분류", "국가", "거래처명", "입금상태",
         "*여기서는 데이터를 변경할 수 없습니다.", "새로고침 하려면 클릭 후 '확인'만 ▼"}


# ─────────────────────── 1. 마스터시트 ───────────────────────
def xl_text(v):
    """xlsx 셀 값을 시트 화면에 보이는 문자열로 되돌린다.
       아래 money/num/pct/kdate 가 전부 그 표기를 기준으로 만들어져 있다."""
    if v is None: return ""
    if isinstance(v, datetime.datetime):
        return f"{v.year % 100}년 {v.month}월 {v.day}일"
    if isinstance(v, datetime.date):
        return f"{v.year % 100}년 {v.month}월 {v.day}일"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def parse_master():
    """마스터시트는 xlsx 원본에서 읽는다.
       텍스트 추출본은 탭당 73행에서 잘려서 절반이 사라진다."""
    wb = openpyxl.load_workbook(MASTER_XLSX, data_only=True)
    ws = wb["26년 SIRIAI 프로젝트"]

    hdr = [xl_text(c.value) for c in ws[6]]
    ix = {h: i for i, h in enumerate(hdr) if h}

    def g(cs, key):
        i = ix.get(key)
        return cs[i] if i is not None and i < len(cs) else ""

    # 시트 자체 대시보드 표기값 (4~5행) — 행 단위 합계와 대조용
    dash_keys = [xl_text(c.value) for c in ws[4] if c.value not in (None, "")]
    dash_vals = [xl_text(c.value) for c in ws[5] if c.value not in (None, "")]
    sheet_dash = dict(zip(dash_keys[1:], dash_vals)) if len(dash_vals) >= 5 else {}

    rows = []
    for r in range(7, ws.max_row + 1):
        rows.append([xl_text(c.value) for c in ws[r]])

    seen, out, anon = set(), [], 0
    for cs in rows:
        if not any(cs): continue
        name = g(cs, "캠페인명")
        junk = (not name) or name in NOISE or "여기서는 데이터를" in name or "모음" in name

        # 이름이 비었어도 금액이 있으면 버리지 않는다 (진행중 건이 그렇게 들어있다)
        if junk:
            if not money(g(cs, "매출(공급가)")):
                continue
            anon += 1
            name = f"(이름 미기재 {anon}) {g(cs,'거래처명') or g(cs,'캠페인 디테일') or ''}".strip()

        if name in seen: continue          # 시트가 탭마다 같은 행을 반복한다
        seen.add(name)

        status_raw = g(cs, "상태")
        status = re.sub(r"^\d+\.\s*", "", status_raw).strip() or "미지정"

        out.append({
            "id": sid("c", name),
            "code": g(cs, "UV") or None,
            "name": name,
            "status": status,
            "month": kmonth(g(cs, "귀속년월")),
            "monthLabel": g(cs, "귀속년월") or None,
            "detail": g(cs, "캠페인 디테일") or None,
            "owner": g(cs, "주체") or None,
            "country": g(cs, "국가") or None,
            "brandName": g(cs, "거래처명") or None,
            "links": {
                "recruit": g(cs, "모집 링크") or None,
                "response": g(cs, "응답 링크") or None,
                "guide": g(cs, "가이드 링크") or None,
                "tracker": g(cs, "캠페인 진행시트") or None,
                "review": g(cs, "콘텐츠 검수 시트") or None,
            },
            "dates": {
                "due":   kdate(g(cs, "납품예정일")),
                "start": kdate(g(cs, "시작일")),
                "end":   kdate(g(cs, "종료일")),
                "endLabel": g(cs, "종료일") or None,
            },
            "counts": {
                "offered":  num(g(cs, "제공 수")),
                "selected": num(g(cs, "선정 수")),
                "uploaded": num(g(cs, "업로드 수")),
                "remaining": num(g(cs, "잔여")),
            },
            "rates": {
                "confirm": pct(g(cs, "컨펌율")),
                "upload":  pct(g(cs, "업로드율")),
            },
            "finance": {
                "expected":   money(g(cs, "입금예정액")),
                "revenue":    money(g(cs, "매출(공급가)")),
                "fee":        money(g(cs, "원고료")),
                "profit":     money(g(cs, "순이익")),
                "margin":     pct(g(cs, "마진율")),
                "quoteDate":  kdate(g(cs, "견적서 발급일")),
                "invoiceDate": kdate(g(cs, "세금계산서 발급일")),
                "paymentStatus": g(cs, "입금상태") or None,
            },
            "note":   g(cs, "비고") or None,
            "remark": g(cs, "특이사항") or None,
        })
    return out, sheet_dash


# ────────────────── 2. 캠페인 진행시트 (참여자) ──────────────────
def parse_participations():
    lines = load(PARTS)
    blocks, i = [], 0
    while i < len(lines):
        cs = cells(lines[i])
        is_header = "인플루언서" in cs and "팔로워" in cs and "No." in cs
        if not is_header:
            i += 1; continue

        # 탭 제목 — 헤더 위쪽에서 '리스트'가 들어간 셀을 찾는다
        title = None
        for j in range(i - 1, max(-1, i - 10), -1):
            for c in cells(lines[j]):
                if "리스트" in c and len(c) > 10:
                    title = c; break
            if title: break

        hx = {h: k for k, h in enumerate(cs)}
        rows, i = [], i + 1
        while i < len(lines):
            r = cells(lines[i])
            if len(r) < 5 or all(x == "" for x in r):     break
            if "인플루언서" in r and "팔로워" in r:        break
            rows.append(r); i += 1
        blocks.append({"title": title, "hx": hx, "rows": rows})

    def base(h):
        """'송장번호(로젠)' · '송장 (롯데)' → 괄호 앞부분만"""
        return h.split("(")[0].strip()

    def g(r, hx, *keys):
        """탭마다 헤더 표기가 다르다. 정확히 맞는 이름을 먼저 쓰고,
           없으면 괄호를 뗀 이름으로 맞춘다. 그 이상은 추측하지 않는다."""
        for key in keys:
            k = hx.get(key)
            if k is not None and k < len(r) and r[k]:
                return r[k]
        bmap = {}
        for h, k in hx.items():
            bmap.setdefault(base(h), k)
        for key in keys:
            k = bmap.get(base(key))
            if k is not None and k < len(r) and r[k]:
                return r[k]
        return ""

    out = []
    for b in blocks:
        title, hx = b["title"] or "제목없음", b["hx"]
        for row_i, r in enumerate(b["rows"]):
            nm = g(r, hx, "인플루언서")
            if not nm or nm in NOISE: continue
            url = norm_url(g(r, hx, "Instagram"))
            sel = g(r, hx, "진행 선정")
            out.append({
                # 같은 탭에 같은 사람이 두 번 오르는 경우가 있어 행 번호까지 넣는다
                "id": sid("p", title, row_i, nm, url or ""),
                "trackerTitle": title,
                "no": g(r, hx, "No.") or None,
                "displayName": nm,
                "handleUrl": url,
                "followers": num(g(r, hx, "팔로워")),
                "region": g(r, hx, "지역") or None,
                "language": g(r, hx, "언어") or None,
                "estViews": num(g(r, hx, "예상 조회수")),
                "er": dec(g(r, hx, "최신10개 게시물 인게이지먼트율")),
                "avgLikes": num(g(r, hx, "최신10개 게시물 평균 좋아요수")),
                "avgComments": num(g(r, hx, "최신10개 게시물 평균 댓글수")),
                "proposedFee": money(g(r, hx, "제안비용")),
                "selected": sel == "선정",
                "campaignChoice": g(r, hx, "캠페인 선택", "제품 옵션") or None,
                "remark": g(r, hx, "비고") or None,
                "managerComment": g(r, hx, "담당자 코멘트", "담당자") or None,
                # 개인정보 — 화면에서는 기본 가림 처리
                "pii": {
                    "realName": g(r, hx, "성함") or None,
                    "phone":    g(r, hx, "연락처") or None,
                    "address":  g(r, hx, "주소지") or None,
                },
                "shipping": g(r, hx, "송장번호", "송장", "소장번호") or None,
                "contentUrl": g(r, hx, "콘텐츠 링크", "콘텐츠 업로드일") or None,
                "threadMirror": g(r, hx, "스레드 미러링") or None,
                "feedback": g(r, hx, "피드백 및 기타") or None,
            })
    return out


# ─────────────────── 3. 데이터 점검 ───────────────────
ADDR = re.compile(r"(시|도|군|구|읍|면|동|로|길)\s|아파트|번길|\d+호")
SHIP_OK = re.compile(r"출고|완료|발송|배송|예정|접수")


def quality(parts):
    """탭마다 열 개수가 달라 값이 옆 칸으로 밀린 행이 있다.
       모양이 확실히 어긋난 값은 비우고, 몇 건이었는지 남긴다."""
    issues = []

    def flag(p, field, val, why):
        issues.append({"field": field, "value": str(val)[:44], "why": why,
                       "tracker": p["trackerTitle"][:40], "who": p["displayName"]})

    for p in parts:
        s = p["shipping"]
        if s:
            if ADDR.search(s) or len(s) > 24:
                flag(p, "송장번호", s, "송장번호 칸에 주소로 보이는 값"); p["shipping"] = None
            elif not re.search(r"\d", s) and not SHIP_OK.search(s):
                flag(p, "송장번호", s, "송장번호로 볼 수 없는 값"); p["shipping"] = None

        ph = p["pii"]["phone"]
        if ph:
            d = re.sub(r"\D", "", ph)
            if not (9 <= len(d) <= 12):
                flag(p, "연락처", ph, "전화번호 형식이 아님"); p["pii"]["phone"] = None

        u = p["handleUrl"]
        if u and "instagram.com/" not in u and "youtube.com/" not in u and "tiktok.com/" not in u:
            flag(p, "계정 주소", u, "알 수 없는 주소 형식"); p["handleUrl"] = None

    return issues


# ─────────────── 4. 진행시트 ↔ 캠페인 연결 ───────────────
WEEK = re.compile(r"(\d{1,2})월\s*(\d)\s*주차")


def match_key(text):
    """'…26년 8월 3주차 (허그유어스킨)' / '[캠페인] 오드타입 26년 8월 3주차' → (월, 주차)"""
    if not text: return None
    m = WEEK.search(text)
    return (int(m.group(1)), int(m.group(2))) if m else None


def link(campaigns, parts):
    by_tracker = defaultdict(list)
    for p in parts:
        by_tracker[p["trackerTitle"]].append(p)

    linked, unlinked = {}, []
    for title, rows in by_tracker.items():
        key = match_key(title)
        brand_hint = title.split("_")[0].strip() if "_" in title else ""
        best = None
        for c in campaigns:
            if key and match_key(c["name"]) == key:
                if brand_hint and brand_hint not in c["name"]:
                    continue
                best = c; break
        if best:
            linked[title] = best["id"]
            for p in rows: p["campaignId"] = best["id"]
        else:
            unlinked.append((title, len(rows)))
            for p in rows: p["campaignId"] = None
    return linked, unlinked


# ─────────────────────────── 실행 ───────────────────────────
# ────────────── 3.5 협업 인플루언서 명단 ──────────────
# 진행시트를 전부 되짚어 "누구와 몇 번 일했는지"를 계정 단위로 모아둔 파일.
# 마스터시트의 협업여부 컬럼은 전 건이 같은 값이라 쓸 수 없어서, 이쪽을 협업 이력의 기준으로 삼는다.

PLATFORM = {
    "instagram": "인스타그램", "tiktok": "틱톡", "youtube": "유튜브",
    "naverblog": "네이버블로그", "naverclip": "네이버클립", "navertv": "네이버TV",
    "navercreator": "네이버", "naver": "네이버", "threads": "스레드",
    "twitter": "X", "x": "X", "x.com": "X",
    "xiaohongshu": "샤오홍슈", "reddit": "레딧", "musinsa": "무신사",
}

CAMPAIGN_KEEP = 8   # 캠페인 목록은 화면에 보여줄 만큼만 남긴다


def norm_handle(url, name):
    """프로필 주소에서 계정만 남긴다. 같은 사람 판별의 기준."""
    s = (url or "").strip()
    for a in ("https://", "http://", "www."):
        s = s.replace(a, "")
    for host in ("instagram.com/", "tiktok.com/", "youtube.com/",
                 "blog.naver.com/", "m.blog.naver.com/", "threads.net/", "threads.com/"):
        s = s.replace(host, "")
    s = s.split("?")[0].strip("/@ ").lower()
    return s or (name or "").strip().lstrip("@").lower()


def parse_creators():
    if not CREATORS_CSV.exists():
        return []
    import csv
    out = []
    with open(CREATORS_CSV, encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            handle = norm_handle(r.get("프로필 링크"), r.get("계정명"))
            if not handle: continue
            raw = (r.get("캠페인 목록") or "").split("|")
            camps = [c.strip() for c in raw if c.strip()]
            plat = (r.get("플랫폼") or "").strip().lower()
            out.append({
                "id": sid("cr", plat + ":" + handle),
                "platform": plat,
                "platformLabel": PLATFORM.get(plat, plat or "기타"),
                "handle": handle,
                "name": (r.get("계정명") or "").strip(),
                "url": (r.get("프로필 링크") or "").strip() or None,
                "campaignCount": num(r.get("참여 캠페인 수")) or 0,
                "appearances": num(r.get("등장 횟수")) or 0,
                "campaigns": camps[:CAMPAIGN_KEEP],
                "campaignsMore": max(0, len(camps) - CAMPAIGN_KEEP),
            })
    out.sort(key=lambda c: (-c["campaignCount"], c["handle"]))
    return out


def main():
    campaigns, sheet_dash = parse_master()
    parts     = parse_participations()
    creators  = parse_creators()
    issues    = quality(parts)
    linked, unlinked = link(campaigns, parts)

    # 브랜드 추출
    bmap = {}
    for c in campaigns:
        bn = c["brandName"]
        if not bn: continue
        bid = bmap.setdefault(bn, {"id": sid("b", bn), "name": bn, "campaignCount": 0})["id"]
        bmap[bn]["campaignCount"] += 1
        c["brandId"] = bid
    brands = sorted(bmap.values(), key=lambda b: -b["campaignCount"])

    # 집계
    def s(key):
        return sum(c["finance"][key] or 0 for c in campaigns)

    by_month = defaultdict(lambda: {"revenue": 0, "profit": 0, "count": 0})
    for c in campaigns:
        m = c["finance"]["invoiceDate"][:7] if c["finance"]["invoiceDate"] else None
        if not m: continue
        by_month[m]["revenue"] += c["finance"]["revenue"] or 0
        by_month[m]["profit"]  += c["finance"]["profit"] or 0
        by_month[m]["count"]   += 1

    revenue, profit = s("revenue"), s("profit")
    paid = sum(c["finance"]["revenue"] or 0 for c in campaigns
               if (c["finance"]["paymentStatus"] or "") == "입금완료")
    uploaded = sum(c["counts"]["uploaded"] or 0 for c in campaigns)
    selected = sum(c["counts"]["selected"] or 0 for c in campaigns)

    db = {
        "meta": {
            "source": "SIRIAI 프로젝트관리 · 무신사 콘텐츠 제작 협의 인플루언서 리스트",
            "note": "운영 구글시트에서 추출한 실데이터입니다.",
            "campaignCount": len(campaigns),
            "participationCount": len(parts),
            "trackerCount": len(set(p["trackerTitle"] for p in parts)),
            "linkedTrackers": len(linked),
            "unlinkedTrackers": [{"title": t, "rows": n} for t, n in unlinked],
            "dataIssueCount": len(issues),
            "dataIssues": issues[:40],
            "creatorCount": len(creators),
            "creatorLinks": sum(c["campaignCount"] for c in creators),
        },
        "kpi": {
            "revenue": revenue,
            "profit": profit,
            "paid": paid,
            "outstanding": revenue - paid,
            "margin": round(profit / revenue, 4) if revenue else 0,
            "uploadRate": round(uploaded / selected, 4) if selected else 0,
            "campaignsTotal": len(campaigns),
            "campaignsRunning": sum(1 for c in campaigns if "종료" not in (c["status"] or "")),
        },
        # 시트가 스스로 표기한 값. 행 단위 합계와 일치하지 않으며, 그 차이를 화면에 드러낸다.
        "sheetDashboard": {
            "revenue": money(sheet_dash.get("총 매출 (공급가)")),
            "profit": money(sheet_dash.get("총 순이익")),
            "paid": money(sheet_dash.get("총 입금 완료액")),
            "outstanding": money(sheet_dash.get("총 미수금액")),
            "margin": pct(sheet_dash.get("당월 마진율")),
            "uploadRate": pct(sheet_dash.get("전체 업로드율")),
            "running": num(sheet_dash.get("진행중 캠페인")),
        },
        "byMonth": [{"month": k, **v} for k, v in sorted(by_month.items())],
        "brands": brands,
        "campaigns": campaigns,
        "participations": parts,
        "creators": creators,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(db, ensure_ascii=False, indent=1), encoding="utf-8")

    print(f"캠페인          {len(campaigns):>5}건")
    print(f"브랜드          {len(brands):>5}개")
    print(f"참여 레코드      {len(parts):>5}건  ({db['meta']['trackerCount']}개 진행시트 탭)")
    print(f"캠페인 연결      {len(linked):>5}/{db['meta']['trackerCount']} 탭")
    print(f"협업 인플루언서  {len(creators):>5}명  (참여 연결 {db['meta']['creatorLinks']:,}건)")
    print(f"누적 매출       ₩{revenue:,}")
    print(f"누적 순이익     ₩{profit:,}   마진 {db['kpi']['margin']*100:.2f}%")
    print(f"미수금         ₩{revenue-paid:,}")
    print(f"데이터 점검      {len(issues):>5}건 걸러냄")
    print(f"\n→ {OUT}  ({OUT.stat().st_size/1024:.0f} KB)")
    if issues:
        kinds = defaultdict(int)
        for it in issues: kinds[f'{it["field"]} — {it["why"]}'] += 1
        for k, v in sorted(kinds.items(), key=lambda x: -x[1]):
            print(f"   {v:>4}건  {k}")
    if unlinked:
        print(f"\n[연결 실패 {len(unlinked)}개 탭 — 화면에서 수동 연결 대상]")
        for t, n in unlinked[:8]:
            print(f"   {n:>3}건  {t[:66]}")


if __name__ == "__main__":
    main()
